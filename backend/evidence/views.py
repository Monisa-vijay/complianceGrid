from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authentication import SessionAuthentication
from rest_framework.views import APIView
from django.utils import timezone
from django.db.models import Q, Count, Prefetch
from django.http import HttpResponse
from datetime import timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import (
    EvidenceCategory, EvidenceSubmission, EvidenceFile,
    SubmissionComment, EvidenceStatus, CategoryGroup, Notification,
    GoogleDriveFolderMapping
)
from .serializers import (
    EvidenceCategorySerializer, EvidenceCategoryDetailSerializer,
    EvidenceSubmissionSerializer, EvidenceFileSerializer,
    SubmissionCommentSerializer, DashboardStatsSerializer, UserSerializer,
    NotificationSerializer
)
from .services.google_drive import GoogleDriveService
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.conf import settings
import requests


def create_due_date_notifications():
    """
    Automatically create notifications for assignees when due date arrives.
    This function checks for submissions with due_date = today and creates notifications
    if they don't already exist. Called automatically from key endpoints.
    """
    today = timezone.now().date()
    submissions = EvidenceSubmission.objects.filter(
        due_date=today,
        status=EvidenceStatus.PENDING
    ).select_related('category', 'category__assignee')
    
    notifications_created = 0
    for submission in submissions:
        category = submission.category
        # Send notification to assignee if exists
        if category.assignee:
            # Check if notification already exists for this submission today
            existing_notification = Notification.objects.filter(
                user=category.assignee,
                submission=submission,
                notification_type='OVERDUE',
                created_at__date=today
            ).first()
            
            if not existing_notification:
                # Create notification with link to control-file page
                Notification.objects.create(
                    user=category.assignee,
                    notification_type='OVERDUE',
                    title=f'Due Today: {category.name}',
                    message=f'Evidence submission for "{category.name}" is due today. Please submit your evidence files.',
                    category=category,
                    submission=submission,
                    is_read=False
                )
                notifications_created += 1
    
    return notifications_created


class EvidenceCategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing evidence categories
    """
    queryset = EvidenceCategory.objects.all()
    permission_classes = []  # AllowAny for development
    # Re-enable pagination with customizable page size
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return EvidenceCategoryDetailSerializer
        return EvidenceCategorySerializer
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def get_queryset(self):
        queryset = EvidenceCategory.objects.all()
        
        # Filter by active status - if showHidden is true, show all; otherwise filter by active_only
        show_hidden = self.request.query_params.get('show_hidden', 'false') == 'true'
        if not show_hidden and self.request.query_params.get('active_only') == 'true':
            queryset = queryset.filter(is_active=True)
        
        # Search by name or description
        search = self.request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(description__icontains=search)
            )
        
        # Filter by review period
        review_period = self.request.query_params.get('review_period', '')
        if review_period:
            queryset = queryset.filter(review_period=review_period)
        
        # Filter by category group
        category_group = self.request.query_params.get('category_group', '')
        if category_group:
            queryset = queryset.filter(category_group=category_group)
        
        # Filter by submission status
        status = self.request.query_params.get('status', '')
        if status:
            if status == 'overdue':
                queryset = queryset.filter(
                    submissions__status='PENDING',
                    submissions__due_date__lt=timezone.now().date()
                ).distinct()
            elif status == 'pending':
                queryset = queryset.filter(
                    submissions__status='PENDING'
                ).distinct()
            elif status == 'submitted':
                queryset = queryset.filter(
                    submissions__status__in=['SUBMITTED', 'UNDER_REVIEW']
                ).distinct()
            elif status == 'approved':
                queryset = queryset.filter(
                    submissions__status='APPROVED'
                ).distinct()
        
        # Ensure proper ordering
        queryset = queryset.order_by('name')
        
        return queryset.prefetch_related(
            'assigned_reviewers', 
            'submissions',
            'submissions__files',
            'submissions__comments'
        ).select_related('primary_assignee', 'assignee', 'approver', 'created_by')
    
    def update(self, request, *args, **kwargs):
        """Override update to send notification when assignee is changed"""
        instance = self.get_object()
        old_assignee = instance.assignee
        
        # Perform the update
        response = super().update(request, *args, **kwargs)
        
        # Check if assignee was changed
        instance.refresh_from_db()
        new_assignee = instance.assignee
        
        if new_assignee and new_assignee != old_assignee:
            # Create notification for the new assignee
            Notification.objects.create(
                user=new_assignee,
                notification_type='CONTROL_ASSIGNED',
                title=f'Control Assigned: {instance.name}',
                message=f'You have been assigned to the control "{instance.name}". Please review and submit evidence as required.',
                category=instance,
                is_read=False
            )
        
        return response
    
    @action(detail=True, methods=['get'])
    def submissions(self, request, pk=None):
        """Get all submissions for a category"""
        category = self.get_object()
        submissions = category.submissions.all().order_by('-due_date')
        serializer = EvidenceSubmissionSerializer(submissions, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def groups(self, request):
        """Get all category groups with counts and compliance scores"""
        show_hidden = request.query_params.get('show_hidden', 'false') == 'true'
        base_queryset = EvidenceCategory.objects.prefetch_related(
            Prefetch(
                'submissions',
                queryset=EvidenceSubmission.objects.prefetch_related('files').select_related('submitted_by', 'reviewed_by')
            )
        ).all()
        if show_hidden:
            # When showing hidden, only show inactive categories
            base_queryset = base_queryset.filter(is_active=False)
        else:
            # When showing active, only show active categories
            base_queryset = base_queryset.filter(is_active=True)
        
        groups = []
        for group_code, group_label in CategoryGroup.choices:
            group_categories = base_queryset.filter(category_group=group_code)
            count = group_categories.count()
            
            if count > 0 or show_hidden:
                # Calculate average compliance score for the group
                total_score = 0
                categories_with_score = 0
                pending_count = 0
                
                for category in group_categories:
                    try:
                        # Check if score should be reset
                        if category.should_reset_compliance_score():
                            score = 0
                        else:
                            score = category.calculate_compliance_score()
                        total_score += score
                        categories_with_score += 1
                        
                        # Count pending evidence (categories with PENDING/REJECTED status and no files, or no submission)
                        # Use prefetched submissions - Django will use prefetched data automatically
                        submissions = category.submissions.all()
                        
                        current_submission = None
                        for sub in submissions:
                            if sub.status in ['PENDING', 'SUBMITTED', 'UNDER_REVIEW', 'REJECTED']:
                                if current_submission is None or sub.due_date > current_submission.due_date:
                                    current_submission = sub
                        
                        if not current_submission:
                            # No active submission - pending
                            pending_count += 1
                        elif current_submission.status in ['PENDING', 'REJECTED']:
                            # Has submission but no files uploaded yet
                            # Check if files are prefetched
                            if hasattr(current_submission, '_prefetched_objects_cache') and 'files' in current_submission._prefetched_objects_cache:
                                files_exist = len(current_submission._prefetched_objects_cache['files']) > 0
                            else:
                                files_exist = current_submission.files.exists()
                            
                            if not files_exist:
                                pending_count += 1
                    except Exception as e:
                        # Log error but continue processing other categories
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error processing category {category.id}: {e}")
                        continue
                
                avg_compliance_score = round(total_score / categories_with_score, 2) if categories_with_score > 0 else 0
                
                groups.append({
                    'code': group_code,
                    'label': group_label,
                    'count': count,
                    'compliance_score': avg_compliance_score,
                    'pending_evidence_count': pending_count
                })
        
        return Response(groups)
    
    @action(detail=False, methods=['post'], url_path='create-google-drive-folders')
    def create_google_drive_folders(self, request):
        """Create folder structure in Google Drive for category groups"""
        # Check if user has Google access token in session
        access_token = request.session.get('google_access_token')
        if not access_token:
            return Response(
                {'error': 'Google Drive authentication required. Please authenticate with Google first.'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        try:
            # Initialize Google Drive service
            refresh_token = request.session.get('google_refresh_token')
            if not access_token:
                return Response(
                    {'error': 'Google Drive not authenticated. Please authenticate Google Drive first.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            drive_service = GoogleDriveService(access_token=access_token, refresh_token=refresh_token)
            
            # Define folder structure mapping
            # Parent categories and their subcategories
            folder_structure = {
                'Security (CC6)': [
                    'ACCESS_CONTROLS',
                    'NETWORK_SECURITY',
                    'PHYSICAL_SECURITY',
                    'DATA_PROTECTION',
                    'ENDPOINT_SECURITY',
                    'MONITORING_INCIDENT'
                ],
                'Availability (CC7)': [
                    'INFRASTRUCTURE_CAPACITY',
                    'BACKUP_RECOVERY',
                    'BUSINESS_CONTINUITY'
                ],
                'Confidentiality (CC8)': [
                    'CONFIDENTIALITY'
                ],
                'Common Criteria (CC1-CC5)': [
                    'CONTROL_ENVIRONMENT',
                    'COMMUNICATION_INFO',
                    'RISK_ASSESSMENT',
                    'MONITORING',
                    'HR_TRAINING',
                    'CHANGE_MANAGEMENT',
                    'VENDOR_MANAGEMENT'
                ]
            }
            
            # Get or create folder mapping record
            folder_mapping, created = GoogleDriveFolderMapping.objects.get_or_create(
                id=1  # Single global mapping
            )
            
            # Create root folder
            if not folder_mapping.root_folder_id:
                root_folder_id = drive_service.create_folder('complianceGrid')
                folder_mapping.root_folder_id = root_folder_id
            else:
                root_folder_id = folder_mapping.root_folder_id
            
            category_group_folder_ids = folder_mapping.category_group_folder_ids or {}
            
            # Create parent category folders and their subcategory folders
            parent_folder_map = {
                'Security (CC6)': 'security_folder_id',
                'Availability (CC7)': 'availability_folder_id',
                'Confidentiality (CC8)': 'confidentiality_folder_id',
                'Common Criteria (CC1-CC5)': 'common_criteria_folder_id'
            }
            
            for parent_name, group_codes in folder_structure.items():
                # Get or create parent folder
                parent_field = parent_folder_map[parent_name]
                parent_folder_id = getattr(folder_mapping, parent_field)
                
                if not parent_folder_id:
                    parent_folder_id = drive_service.create_folder(parent_name, parent_folder_id=root_folder_id)
                    setattr(folder_mapping, parent_field, parent_folder_id)
                
                # Create subcategory folders (category group folders)
                for group_code in group_codes:
                    if group_code not in category_group_folder_ids:
                        # Get the label for this group
                        group_label = dict(CategoryGroup.choices).get(group_code, group_code)
                        subfolder_id = drive_service.create_folder(group_label, parent_folder_id=parent_folder_id)
                        category_group_folder_ids[group_code] = subfolder_id
            
            # Save folder mapping
            folder_mapping.category_group_folder_ids = category_group_folder_ids
            folder_mapping.save()
            
            # Step 2: Create folders for each EvidenceCategory (control) inside their category group folders
            categories_created = 0
            categories_skipped = 0
            for group_code, group_folder_id in category_group_folder_ids.items():
                # Get all categories for this group
                categories = EvidenceCategory.objects.filter(
                    category_group=group_code,
                    is_active=True
                )
                
                for category in categories:
                    # Skip if folder already exists
                    if category.google_drive_folder_id:
                        categories_skipped += 1
                        continue
                    
                    # Create folder for this category (control)
                    try:
                        category_folder_id = drive_service.create_folder(
                            category.name,
                            parent_folder_id=group_folder_id
                        )
                        # Store folder ID in category
                        category.google_drive_folder_id = category_folder_id
                        category.save()
                        categories_created += 1
                    except Exception as e:
                        # Log error but continue with other categories
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error creating folder for category {category.name}: {str(e)}")
                        continue
            
            # Step 3: Sync files - Upload any approved files that haven't been uploaded to Google Drive yet
            files_uploaded = 0
            files_skipped = 0
            files_failed = 0
            upload_errors = []
            
            # Get all approved submissions with files that haven't been uploaded
            approved_submissions = EvidenceSubmission.objects.filter(
                status=EvidenceStatus.APPROVED
            ).prefetch_related('files', 'category')
            
            for submission in approved_submissions:
                category = submission.category
                # Skip if category doesn't have a Google Drive folder ID
                if not category.google_drive_folder_id:
                    continue
                
                # Get files that haven't been uploaded to Google Drive yet
                files_to_upload = submission.files.filter(google_drive_file_id__isnull=True)
                
                for evidence_file in files_to_upload:
                    if evidence_file.file:  # Check if local file exists
                        try:
                            # Read file content
                            evidence_file.file.open('rb')
                            file_content = evidence_file.file.read()
                            evidence_file.file.close()
                            
                            # Upload to Google Drive
                            drive_result = drive_service.upload_file(
                                file_content=file_content,
                                filename=evidence_file.filename,
                                folder_id=category.google_drive_folder_id,
                                mime_type=evidence_file.mime_type
                            )
                            
                            # Store Google Drive file info
                            evidence_file.google_drive_file_id = drive_result['file_id']
                            evidence_file.google_drive_file_url = drive_result['web_url']
                            evidence_file.save()
                            files_uploaded += 1
                        except Exception as e:
                            # Log error but continue with other files
                            import logging
                            logger = logging.getLogger(__name__)
                            error_msg = f"Failed to upload {evidence_file.filename} to Google Drive: {str(e)}"
                            logger.error(error_msg)
                            upload_errors.append(error_msg)
                            files_failed += 1
                    else:
                        error_msg = f"Local file not found for {evidence_file.filename}"
                        upload_errors.append(error_msg)
                        files_failed += 1
            
            # Build response message
            message_parts = ['Folder structure synced successfully']
            if categories_created > 0:
                message_parts.append(f'{categories_created} category folder(s) created')
            if files_uploaded > 0:
                message_parts.append(f'{files_uploaded} file(s) uploaded to Google Drive')
            if files_failed > 0:
                message_parts.append(f'{files_failed} file(s) failed to upload')
            
            response_data = {
                'message': '. '.join(message_parts) + '.',
                'root_folder_id': root_folder_id,
                'categories_created': categories_created,
                'categories_skipped': categories_skipped,
                'files_uploaded': files_uploaded,
                'files_failed': files_failed,
                'folder_mapping': {
                    'security_folder_id': folder_mapping.security_folder_id,
                    'availability_folder_id': folder_mapping.availability_folder_id,
                    'confidentiality_folder_id': folder_mapping.confidentiality_folder_id,
                    'common_criteria_folder_id': folder_mapping.common_criteria_folder_id,
                    'category_group_folder_ids': category_group_folder_ids
                }
            }
            
            if upload_errors:
                response_data['upload_errors'] = upload_errors
            
            return Response(response_data)
            
        except Exception as e:
            import traceback
            return Response(
                {'error': f'Failed to create folder structure: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='users')
    def get_users(self, request):
        """Get all users for assignee/approver selection"""
        users = User.objects.all().order_by('username')
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_groups(self, request):
        """Export category groups data in PDF or Excel format"""
        try:
            format_type = request.query_params.get('format', 'excel').lower()
            show_hidden = request.query_params.get('show_hidden', 'false') == 'true'
            
            # Get all categories with their submissions and files
            base_queryset = EvidenceCategory.objects.select_related(
                'assignee', 'approver'
            ).prefetch_related(
                Prefetch(
                    'submissions',
                    queryset=EvidenceSubmission.objects.prefetch_related(
                        'files', 'files__uploaded_by'
                    ).select_related('submitted_by', 'reviewed_by')
                )
            ).all()
            
            if show_hidden:
                # When showing hidden, only show inactive categories
                base_queryset = base_queryset.filter(is_active=False)
            else:
                # When showing active, only show active categories
                base_queryset = base_queryset.filter(is_active=True)
            
            # Prepare export data
            export_data = []
            for group_code, group_label in CategoryGroup.choices:
                if group_code == 'UNCATEGORIZED':
                    continue
                    
                group_categories = base_queryset.filter(category_group=group_code)
                
                for category in group_categories:
                    try:
                        # Get all submissions for this category (using prefetched data)
                        submissions = list(category.submissions.all())
                        
                        # Get latest submission with files
                        latest_submission = None
                        latest_file = None
                        for sub in submissions:
                            # Check if submission has files (using prefetched data)
                            sub_files = list(sub.files.all())
                            if sub_files:
                                # Find the latest file in this submission
                                sub_latest_file = None
                                for f in sub_files:
                                    if sub_latest_file is None or (f.uploaded_at and sub_latest_file.uploaded_at and f.uploaded_at > sub_latest_file.uploaded_at):
                                        sub_latest_file = f
                                
                                if sub_latest_file:
                                    if latest_submission is None or (sub.submitted_at and latest_submission.submitted_at and sub.submitted_at > latest_submission.submitted_at):
                                        latest_submission = sub
                                        latest_file = sub_latest_file
                        
                        # Get file details
                        uploaded_by = None
                        uploaded_date = None
                        approved_by = None
                        
                        if latest_file:
                            uploaded_by = latest_file.uploaded_by.username if latest_file.uploaded_by else 'N/A'
                            uploaded_date = latest_file.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if latest_file.uploaded_at else 'N/A'
                        
                        if latest_submission and latest_submission.status == 'APPROVED' and latest_submission.reviewed_by:
                            approved_by = latest_submission.reviewed_by.username
                        
                        # Determine evidence status
                        current_submission = None
                        for sub in submissions:
                            if sub.status in [EvidenceStatus.PENDING, EvidenceStatus.SUBMITTED, 
                                             EvidenceStatus.UNDER_REVIEW, EvidenceStatus.REJECTED]:
                                if current_submission is None or sub.due_date > current_submission.due_date:
                                    current_submission = sub
                        
                        if not current_submission:
                            evidence_status = 'Missing'
                        elif current_submission.status in [EvidenceStatus.PENDING, EvidenceStatus.REJECTED]:
                            # Check if files exist using prefetched data
                            current_files = list(current_submission.files.all())
                            if not current_files:
                                evidence_status = 'Missing'
                            else:
                                evidence_status = 'Uploaded'
                        else:
                            # Check if files exist using prefetched data
                            current_files = list(current_submission.files.all())
                            if current_files:
                                evidence_status = 'Uploaded'
                            else:
                                evidence_status = 'Missing'
                        
                        export_data.append({
                            'category_group': group_label,
                            'control': category.name,
                            'evidence_status': evidence_status,
                            'last_uploaded_date': uploaded_date or 'N/A',
                            'uploaded_by': uploaded_by or 'N/A',
                            'approved_by': approved_by or 'N/A'
                        })
                    except Exception as e:
                        # Log error but continue with other categories
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error processing category {category.id} for export: {e}")
                        # Still add the category with default values
                        export_data.append({
                            'category_group': group_label,
                            'control': category.name,
                            'evidence_status': 'Error',
                            'last_uploaded_date': 'N/A',
                            'uploaded_by': 'N/A',
                            'approved_by': 'N/A'
                        })
                        continue
            
            # Check if we have data to export
            if not export_data:
                return Response(
                    {'error': 'No data available to export'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            if format_type == 'pdf':
                return self._generate_pdf(export_data)
            else:
                return self._generate_excel(export_data)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in export_groups: {e}", exc_info=True)
            return Response(
                {'error': f'Failed to export data: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_excel(self, data):
        """Generate Excel file"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Category Groups Export"
            
            # Headers
            headers = ['Category Group', 'Control', 'Evidence Status', 'Last Uploaded Date', 'Uploaded By', 'Approved By']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Add data
            for row in data:
                ws.append([
                    row['category_group'],
                    row['control'],
                    row['evidence_status'],
                    row['last_uploaded_date'],
                    row['uploaded_by'],
                    row['approved_by']
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="category_groups_export.xlsx"'
            return response
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error generating Excel: {e}", exc_info=True)
            raise
    
    def _generate_pdf(self, data):
        """Generate PDF file"""
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Title
            title = Paragraph("Category Groups Export Report", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Prepare table data
            table_data = [['Category Group', 'Control', 'Evidence Status', 'Last Uploaded Date', 'Uploaded By', 'Approved By']]
            
            for row in data:
                table_data.append([
                    row['category_group'],
                    row['control'],
                    row['evidence_status'],
                    row['last_uploaded_date'],
                    row['uploaded_by'],
                    row['approved_by']
                ])
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="category_groups_export.pdf"'
            return response
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error generating PDF: {e}", exc_info=True)
            raise


class EvidenceSubmissionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing evidence submissions
    """
    queryset = EvidenceSubmission.objects.all()
    serializer_class = EvidenceSubmissionSerializer
    permission_classes = []  # AllowAny for development
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def get_queryset(self):
        queryset = EvidenceSubmission.objects.select_related(
            'category', 'submitted_by', 'reviewed_by'
        ).prefetch_related('files', 'comments')
        
        # Filter by category if provided
        category_id = self.request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        # Filter by status if provided
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit evidence files for a submission"""
        submission = self.get_object()
        
        if submission.status not in [EvidenceStatus.PENDING, EvidenceStatus.REJECTED]:
            return Response(
                {'error': 'This submission cannot be submitted in its current status.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate that category has both assignee and approver
        category = submission.category
        if not category.assignee:
            return Response(
                {'error': 'Cannot submit files: Assignee is required for this control.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not category.approver:
            return Response(
                {'error': 'Cannot submit files: Approver is required for this control.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get files from request
        files = request.FILES.getlist('files')
        if not files:
            return Response(
                {'error': 'No files provided.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get notes and due date
        notes = request.data.get('notes', '')
        due_date_str = request.data.get('due_date')
        
        # Save files locally (Google Drive upload will happen after approval)
        try:
            uploaded_files = []
            
            for file in files:
                # Create EvidenceFile record with local file storage only
                evidence_file = EvidenceFile.objects.create(
                    submission=submission,
                    filename=file.name,
                    file=file,  # Save file locally using FileField
                    file_size=file.size,
                    mime_type=file.content_type or 'application/octet-stream',
                    uploaded_by=request.user if request.user.is_authenticated else None
                )
                
                uploaded_files.append(evidence_file)
            
            # Update submission
            submission.status = EvidenceStatus.SUBMITTED
            submission.submitted_by = request.user if request.user.is_authenticated else None
            submission.submitted_at = timezone.now()
            submission.submission_notes = notes
            
            # Update due date if provided
            if due_date_str:
                try:
                    from datetime import datetime
                    due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                    submission.due_date = due_date
                except ValueError:
                    pass  # Keep existing due date if parsing fails
            
            submission.save()
            
            # Send notification to approver
            if category.approver:
                Notification.objects.create(
                    user=category.approver,
                    notification_type='PENDING_APPROVAL',
                    title=f'Pending Approval: {category.name}',
                    message=f'New evidence files have been submitted for "{category.name}" and are awaiting your approval.',
                    category=category,
                    submission=submission,
                    is_read=False
                )
            
            serializer = EvidenceSubmissionSerializer(submission, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to upload files: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a submission and upload files to Google Drive"""
        submission = self.get_object()
        
        if submission.status != EvidenceStatus.SUBMITTED and submission.status != EvidenceStatus.UNDER_REVIEW:
            return Response(
                {'error': 'Only submitted or under review submissions can be approved.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        review_notes = request.data.get('review_notes', '')
        
        # Update submission status
        submission.status = EvidenceStatus.APPROVED
        submission.reviewed_by = request.user if request.user.is_authenticated else None
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Upload files to Google Drive after approval
        category = submission.category
        upload_errors = []
        uploaded_count = 0
        
        if category.google_drive_folder_id:
            access_token = request.session.get('google_access_token')
            refresh_token = request.session.get('google_refresh_token')
            
            if not access_token:
                # Try to get token from any user's session (since Google Drive is shared)
                # Check if there's a way to get a shared token or use the first authenticated user
                from django.contrib.sessions.models import Session
                
                # Try to find a recent session with Google Drive token
                # Look for sessions from the last 24 hours
                recent_sessions = Session.objects.filter(expire_date__gte=timezone.now())
                for session in recent_sessions:
                    try:
                        session_data = session.get_decoded()
                        if 'google_access_token' in session_data:
                            access_token = session_data['google_access_token']
                            refresh_token = session_data.get('google_refresh_token')
                            break
                    except Exception:
                        continue
            
            if access_token:
                try:
                    drive_service = GoogleDriveService(access_token=access_token, refresh_token=refresh_token)
                    # Get all files for this submission that haven't been uploaded to Google Drive yet
                    files_to_upload = submission.files.filter(google_drive_file_id__isnull=True)
                    
                    if not files_to_upload.exists():
                        # All files already uploaded
                        serializer = EvidenceSubmissionSerializer(submission)
                        return Response({
                            **serializer.data,
                            'message': 'Submission approved. All files were already uploaded to Google Drive.'
                        })
                    
                    for evidence_file in files_to_upload:
                        if evidence_file.file:  # Check if local file exists
                            try:
                                # Read file content
                                evidence_file.file.open('rb')
                                file_content = evidence_file.file.read()
                                evidence_file.file.close()
                                
                                # Upload to Google Drive
                                drive_result = drive_service.upload_file(
                                    file_content=file_content,
                                    filename=evidence_file.filename,
                                    folder_id=category.google_drive_folder_id,
                                    mime_type=evidence_file.mime_type
                                )
                                
                                # Store Google Drive file info
                                evidence_file.google_drive_file_id = drive_result['file_id']
                                evidence_file.google_drive_file_url = drive_result['web_url']
                                evidence_file.save()
                                uploaded_count += 1
                            except Exception as e:
                                # Log error and collect for response
                                import logging
                                logger = logging.getLogger(__name__)
                                error_msg = f"Failed to upload {evidence_file.filename} to Google Drive: {str(e)}"
                                logger.error(error_msg)
                                upload_errors.append(error_msg)
                        else:
                            error_msg = f"Local file not found for {evidence_file.filename}"
                            upload_errors.append(error_msg)
                            
                except Exception as e:
                    # Log error and collect for response
                    import logging
                    logger = logging.getLogger(__name__)
                    error_msg = f"Failed to initialize Google Drive service: {str(e)}"
                    logger.error(error_msg)
                    upload_errors.append(error_msg)
            else:
                upload_errors.append("Google Drive not authenticated. Please authenticate Google Drive first.")
        else:
            upload_errors.append("Google Drive folder not configured for this category.")
        
        serializer = EvidenceSubmissionSerializer(submission)
        response_data = serializer.data
        
        # Add upload status to response
        if uploaded_count > 0:
            response_data['upload_status'] = f'Successfully uploaded {uploaded_count} file(s) to Google Drive.'
        if upload_errors:
            response_data['upload_errors'] = upload_errors
            response_data['upload_warning'] = 'Submission approved, but some files could not be uploaded to Google Drive.'
        
        return Response(response_data)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a submission"""
        submission = self.get_object()
        
        if submission.status != EvidenceStatus.SUBMITTED and submission.status != EvidenceStatus.UNDER_REVIEW:
            return Response(
                {'error': 'Only submitted or under review submissions can be rejected.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        review_notes = request.data.get('review_notes', '')
        if not review_notes:
            return Response(
                {'error': 'Review notes are required for rejection.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        submission.status = EvidenceStatus.REJECTED
        submission.reviewed_by = request.user if request.user.is_authenticated else None
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        serializer = EvidenceSubmissionSerializer(submission)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get dashboard statistics with gap analysis"""
        # Automatically create due date notifications
        create_due_date_notifications()
        
        today = timezone.now().date()
        start_of_month = today.replace(day=1)
        
        active_categories = EvidenceCategory.objects.filter(is_active=True)
        total_categories = active_categories.count()
        
        # Basic stats
        pending_submissions = EvidenceSubmission.objects.filter(
            status=EvidenceStatus.PENDING
        ).count()
        
        overdue_submissions = EvidenceSubmission.objects.filter(
            status=EvidenceStatus.PENDING,
            due_date__lt=today
        ).count()
        
        approved_this_month = EvidenceSubmission.objects.filter(
            status=EvidenceStatus.APPROVED,
            reviewed_at__gte=start_of_month
        ).count()
        
        # Gap Analysis Metrics
        # Controls without any submissions
        categories_with_submissions = EvidenceSubmission.objects.values_list('category_id', flat=True).distinct()
        controls_without_evidence = active_categories.exclude(id__in=categories_with_submissions).count()
        
        # Controls without assignees
        controls_without_assignee = active_categories.filter(assignee__isnull=True).count()
        
        # Controls without approvers
        controls_without_approver = active_categories.filter(approver__isnull=True).count()
        
        # Controls with overdue submissions
        overdue_category_ids = EvidenceSubmission.objects.filter(
            status=EvidenceStatus.PENDING,
            due_date__lt=today
        ).values_list('category_id', flat=True).distinct()
        controls_with_overdue = active_categories.filter(id__in=overdue_category_ids).count()
        
        # Controls with low compliance (below 50%)
        # Calculate compliance scores efficiently
        controls_with_low_compliance = 0
        categories_with_scores = active_categories.prefetch_related('submissions__files')
        for category in categories_with_scores:
            score = category.calculate_compliance_score()
            if score < 50:
                controls_with_low_compliance += 1
        
        # Controls pending approval
        pending_approval_submissions = EvidenceSubmission.objects.filter(
            status__in=[EvidenceStatus.SUBMITTED, EvidenceStatus.UNDER_REVIEW]
        )
        controls_pending_approval = pending_approval_submissions.values_list('category_id', flat=True).distinct().count()
        
        stats = {
            'total_categories': total_categories,
            'pending_submissions': pending_submissions,
            'overdue_submissions': overdue_submissions,
            'approved_this_month': approved_this_month,
            # Gap Analysis
            'controls_without_evidence': controls_without_evidence,
            'controls_without_assignee': controls_without_assignee,
            'controls_without_approver': controls_without_approver,
            'controls_with_overdue': controls_with_overdue,
            'controls_with_low_compliance': controls_with_low_compliance,
            'controls_pending_approval': controls_pending_approval,
        }
        
        # Get upcoming deadlines (next 30 days)
        upcoming_deadlines = EvidenceSubmission.objects.filter(
            status=EvidenceStatus.PENDING,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=30)
        ).select_related('category').order_by('due_date')[:10]
        
        serializer = DashboardStatsSerializer({
            **stats,
            'upcoming_deadlines': upcoming_deadlines
        })
        
        return Response(serializer.data)


# CSRF-exempt login view using APIView
class LoginView(APIView):
    """Handle email/password login - CSRF exempt"""
    authentication_classes = []
    permission_classes = []
    
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def post(self, request):
        from django.contrib.auth import authenticate, login
        from django.contrib.auth.models import User
        
        email = request.data.get('email') or request.data.get('username')
        password = request.data.get('password')
        
        if not email or not password:
            return Response(
                {'error': 'Email and password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Try to find user by email first
        try:
            users = User.objects.filter(email=email)
            if users.exists():
                # If multiple users with same email, use the first one
                user = users.first()
                username = user.username
            else:
                # Fallback to username if email not found
                try:
                    user = User.objects.get(username=email)
                    username = user.username
                except User.DoesNotExist:
                    return Response(
                        {'error': 'Invalid email or password'},
                        status=status.HTTP_401_UNAUTHORIZED
                    )
        except Exception:
            # Fallback to username if any error
            try:
                user = User.objects.get(username=email)
                username = user.username
            except User.DoesNotExist:
                return Response(
                    {'error': 'Invalid email or password'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Ensure session exists before login
            if not request.session.session_key:
                request.session.create()
            
            # Login the user (this sets request.user and stores auth data in session)
            login(request, user)
            
            # Force session to save with modified flag
            request.session.modified = True
            request.session.save()
            
            # Verify the session has the auth data
            from django.contrib.sessions.models import Session
            session_obj = Session.objects.get(session_key=request.session.session_key)
            print(f"DEBUG: Session data after login: {session_obj.get_decoded()}")
            
            # Get session key after save
            session_key = request.session.session_key
            print(f"DEBUG: Login successful for user {user.username}, session_key: {session_key}, user_id in session: {request.session.get('_auth_user_id')}")
            
            # Set CSRF token in response for subsequent requests
            from django.middleware.csrf import get_token
            csrf_token = get_token(request)
            serializer = UserSerializer(user)
            response = Response({
                'user': serializer.data,
                'message': 'Login successful',
                'session_key': session_key  # Debug: include session key in response
            })
            # Ensure CSRF cookie is set
            response.set_cookie(
                'csrftoken',
                csrf_token,
                max_age=60 * 60 * 24 * 7,  # 7 days
                httponly=False,
                samesite='Lax',
                secure=False,
                path='/',
                domain=None  # Allow cookie for localhost
            )
            # Explicitly set session cookie
            if session_key:
                response.set_cookie(
                    'sessionid',
                    session_key,
                    max_age=60 * 60 * 24 * 7,  # 7 days
                    httponly=True,
                    samesite='Lax',
                    secure=False,
                    path='/',
                    domain=None  # Allow cookie for localhost
                )
                print(f"DEBUG: Setting session cookie: {session_key}")
            else:
                print("DEBUG: WARNING - No session key after login!")
            return response
        else:
            return Response(
                {'error': 'Invalid email or password'},
                status=status.HTTP_401_UNAUTHORIZED
            )


class AuthView(viewsets.ViewSet):
    """
    ViewSet for username/password authentication
    """
    permission_classes = []
    authentication_classes = []  # No authentication required for login
    
    @action(detail=False, methods=['post'])
    def logout(self, request):
        """Handle logout"""
        from django.contrib.auth import logout
        
        try:
            logout(request)
            response = Response({'message': 'Logout successful'})
            # Clear session cookie
            response.delete_cookie('sessionid')
            response.delete_cookie('csrftoken')
            return response
        except Exception as e:
            # Even if logout fails, try to clear cookies
            response = Response({'message': 'Logout completed'}, status=status.HTTP_200_OK)
            response.delete_cookie('sessionid')
            response.delete_cookie('csrftoken')
            return response
    
    @action(detail=False, methods=['get'], authentication_classes=[], permission_classes=[])
    def me(self, request):
        """Get current user - allow unauthenticated to return None"""
        session_key = request.session.session_key
        auth_user_id = request.session.get('_auth_user_id') if session_key else None
        
        # Debug: Check session data from database
        if session_key:
            from django.contrib.sessions.models import Session
            try:
                session_obj = Session.objects.get(session_key=session_key)
                session_data = session_obj.get_decoded()
                print(f"DEBUG: Session data from DB: {session_data}")
                print(f"DEBUG: Auth user ID in DB session: {session_data.get('_auth_user_id')}")
            except Session.DoesNotExist:
                print(f"DEBUG: WARNING - Session {session_key} not found in database!")
        
        print(f"DEBUG: /auth/me/ called - session_key: {session_key}, auth_user_id in session: {auth_user_id}, user: {request.user}, authenticated: {request.user.is_authenticated}")
        print(f"DEBUG: Cookies in request: {request.COOKIES}")
        
        # Check Google Drive authentication status
        google_drive_authenticated = bool(request.session.get('google_access_token'))
        
        # If we have auth_user_id in session but user is not authenticated, try to get user manually
        if not request.user.is_authenticated and auth_user_id:
            from django.contrib.auth.models import User
            try:
                user = User.objects.get(pk=auth_user_id)
                print(f"DEBUG: Found user from session auth_user_id: {user.username}")
                serializer = UserSerializer(user)
                return Response({
                    'user': serializer.data,
                    'google_drive_authenticated': google_drive_authenticated
                })
            except User.DoesNotExist:
                print(f"DEBUG: User with id {auth_user_id} does not exist")
        
        if request.user.is_authenticated:
            serializer = UserSerializer(request.user)
            return Response({
                'user': serializer.data,
                'google_drive_authenticated': google_drive_authenticated
            })
        else:
            return Response({
                'user': None,
                'google_drive_authenticated': False
            }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def csrf(self, request):
        """Get CSRF token and ensure cookie is set"""
        from django.middleware.csrf import get_token
        
        # This ensures the CSRF cookie is set
        csrf_token = get_token(request)
        response = Response({'csrfToken': csrf_token})
        # Explicitly set the cookie in the response
        response.set_cookie(
            'csrftoken',
            csrf_token,
            max_age=60 * 60 * 24 * 7,  # 7 days
            httponly=False,
            samesite='Lax',
            secure=False
        )
        return response
    
    @action(detail=False, methods=['post'])
    def change_password(self, request):
        """Change user password"""
        from django.contrib.auth import update_session_auth_hash
        
        if not request.user.is_authenticated:
            return Response(
                {'error': 'Authentication required'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        
        if not current_password or not new_password:
            return Response(
                {'error': 'Current password and new password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify current password
        if not request.user.check_password(current_password):
            return Response(
                {'error': 'Current password is incorrect'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate new password length
        if len(new_password) < 8:
            return Response(
                {'error': 'New password must be at least 8 characters long'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Set new password
        request.user.set_password(new_password)
        request.user.save()
        
        # Update session to prevent logout
        update_session_auth_hash(request, request.user)
        
        return Response({'message': 'Password changed successfully'})
    
    @action(detail=False, methods=['patch'])
    def update_profile(self, request):
        """Update user profile information"""
        if not request.user.is_authenticated:
            return Response(
                {'error': 'Authentication required'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # Get updatable fields
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        email = request.data.get('email')
        
        # Update fields if provided
        if first_name is not None:
            request.user.first_name = first_name
        if last_name is not None:
            request.user.last_name = last_name
        if email is not None:
            # Validate email format
            from django.core.validators import validate_email
            from django.core.exceptions import ValidationError
            try:
                validate_email(email)
                request.user.email = email
            except ValidationError:
                return Response(
                    {'error': 'Invalid email format'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        request.user.save()
        serializer = UserSerializer(request.user)
        return Response({'user': serializer.data, 'message': 'Profile updated successfully'})


class GoogleAuthView(viewsets.ViewSet):
    """
    ViewSet for Google OAuth authentication (deprecated - kept for backward compatibility)
    """
    permission_classes = []
    
    @action(detail=False, methods=['post'])
    def google(self, request):
        """Handle Google OAuth callback and create/update user"""
        from django.contrib.auth.models import User
        from django.contrib.auth import login
        
        access_token = request.data.get('access_token')
        user_info = request.data.get('user_info', {})
        
        if not access_token or not user_info:
            return Response(
                {'error': 'Missing access token or user info'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get or create user
        email = user_info.get('email', '')
        google_id = user_info.get('sub', '')
        
        if not email:
            return Response(
                {'error': 'Email not provided by Google'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create or get user
        user, created = User.objects.get_or_create(
            username=email,
            defaults={
                'email': email,
                'first_name': user_info.get('given_name', ''),
                'last_name': user_info.get('family_name', ''),
            }
        )
        
        if not created:
            # Update existing user
            user.email = email
            user.first_name = user_info.get('given_name', '')
            user.last_name = user_info.get('family_name', '')
            user.save()
        
        # Store Google access token in session
        request.session['google_access_token'] = access_token
        request.session['google_user_info'] = user_info
        
        # Login user
        login(request, user)
        
        return Response({
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'name': user.get_full_name() or user.username,
            },
            'message': 'Authentication successful'
        })
    
    @action(detail=False, methods=['get'])
    def initiate(self, request):
        """Get Google OAuth authorization URL"""
        from django.conf import settings
        from urllib.parse import urlencode
        
        # Build authorization URL
        params = {
            'client_id': settings.GOOGLE_DRIVE_CLIENT_ID,
            'redirect_uri': settings.GOOGLE_DRIVE_REDIRECT_URI,
            'response_type': 'code',
            'scope': ' '.join(settings.GOOGLE_DRIVE_SCOPES),
            'access_type': 'offline',
            'prompt': 'consent',
        }
        
        auth_url = f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}"
        
        return Response({
            'authorization_url': auth_url
        })


class GoogleOAuthCallbackView(APIView):
    """Handle Google OAuth callback - CSRF exempt"""
    authentication_classes = []
    permission_classes = []
    
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def post(self, request):
        """Handle Google OAuth callback - exchange code for token"""
        from django.contrib.auth.models import User
        from django.contrib.auth import login
        from django.conf import settings
        import requests
        
        code = request.data.get('code')
        if not code:
            return Response(
                {'error': 'Authorization code not provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Exchange code for token
            # Use the redirect_uri from settings (must match the one used in authorization request)
            token_url = 'https://oauth2.googleapis.com/token'
            
            token_data = {
                'code': code,
                'client_id': settings.GOOGLE_DRIVE_CLIENT_ID,
                'client_secret': settings.GOOGLE_DRIVE_CLIENT_SECRET,
                'redirect_uri': settings.GOOGLE_DRIVE_REDIRECT_URI,  # Must match the redirect_uri used in authorization
                'grant_type': 'authorization_code',
            }
            
            token_response = requests.post(token_url, data=token_data)
            token_response.raise_for_status()
            token_json = token_response.json()
            
            access_token = token_json.get('access_token')
            refresh_token = token_json.get('refresh_token')
            
            if not access_token:
                return Response(
                    {'error': 'Failed to get access token'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get user info from Google
            user_info_response = requests.get(
                'https://www.googleapis.com/oauth2/v3/userinfo',
                headers={'Authorization': f'Bearer {access_token}'}
            )
            user_info_response.raise_for_status()
            user_info = user_info_response.json()
            
            # Get or create user
            email = user_info.get('email', '')
            if not email:
                return Response(
                    {'error': 'Email not provided by Google'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create or get user - check by email first to avoid duplicates
            user = None
            try:
                # Try to find user by email first
                user = User.objects.filter(email=email).first()
                if user:
                    # Update existing user
                    user.first_name = user_info.get('given_name', '')
                    user.last_name = user_info.get('family_name', '')
                    user.save()
            except Exception:
                pass
            
            # If no user found by email, try to get or create by username
            if not user:
                user, created = User.objects.get_or_create(
                    username=email,
                    defaults={
                        'email': email,
                        'first_name': user_info.get('given_name', ''),
                        'last_name': user_info.get('family_name', ''),
                    }
                )
                
                if not created:
                    # Update existing user
                    user.email = email
                    user.first_name = user_info.get('given_name', '')
                    user.last_name = user_info.get('family_name', '')
                    user.save()
            
            # Store Google access token in session
            request.session['google_access_token'] = access_token
            if refresh_token:
                request.session['google_refresh_token'] = refresh_token
            request.session['google_user_info'] = user_info
            request.session.modified = True  # Ensure session is saved
            
            # If user is not already logged in, log them in
            # Otherwise, just store the token for the existing session
            if not request.user.is_authenticated:
                login(request, user)
            
            # Return current user (either newly logged in or existing)
            current_user = request.user if request.user.is_authenticated else user
            
            # Set CSRF token in response for subsequent requests
            from django.middleware.csrf import get_token
            csrf_token = get_token(request)
            
            response = Response({
                'user': {
                    'id': current_user.id,
                    'username': current_user.username,
                    'email': current_user.email,
                    'name': current_user.get_full_name() or current_user.username,
                },
                'access_token': access_token,
                'message': 'Google Drive authentication successful'
            })
            
            # Ensure CSRF cookie is set for subsequent requests
            response.set_cookie(
                'csrftoken',
                csrf_token,
                max_age=60 * 60 * 24 * 7,  # 7 days
                httponly=False,
                samesite='Lax',
                secure=False,
                path='/',
                domain=None  # Allow cookie for localhost
            )
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Authentication failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EvidenceFileViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing evidence files/documents
    """
    queryset = EvidenceFile.objects.select_related('submission__category', 'uploaded_by').all()
    serializer_class = EvidenceFileSerializer
    permission_classes = []  # AllowAny for development
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def get_queryset(self):
        queryset = EvidenceFile.objects.select_related('submission__category', 'uploaded_by').all()
        
        # Filter by uploaded_by user
        uploaded_by = self.request.query_params.get('uploaded_by')
        if uploaded_by:
            queryset = queryset.filter(uploaded_by_id=uploaded_by)
        
        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(submission__category_id=category)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(uploaded_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(uploaded_at__date__lte=date_to)
        
        return queryset.order_by('-uploaded_at')
    
    @action(detail=False, methods=['get'])
    def grouped(self, request):
        """Get all documents grouped by uploaded date and uploaded by"""
        files = self.get_queryset()
        
        # Group by date and user
        grouped_data = {}
        for file in files:
            date_key = file.uploaded_at.date().isoformat()
            user_key = file.uploaded_by.username if file.uploaded_by else 'Unknown'
            
            if date_key not in grouped_data:
                grouped_data[date_key] = {}
            
            if user_key not in grouped_data[date_key]:
                grouped_data[date_key][user_key] = {
                    'user': {
                        'id': file.uploaded_by.id if file.uploaded_by else None,
                        'username': user_key,
                        'email': file.uploaded_by.email if file.uploaded_by else None,
                    },
                    'files': []
                }
            
            serializer = EvidenceFileSerializer(file, context={'request': request})
            grouped_data[date_key][user_key]['files'].append(serializer.data)
        
        # Convert to list format sorted by date (newest first)
        result = []
        for date_key in sorted(grouped_data.keys(), reverse=True):
            date_entry = {
                'date': date_key,
                'users': []
            }
            for user_key in sorted(grouped_data[date_key].keys()):
                date_entry['users'].append(grouped_data[date_key][user_key])
            result.append(date_entry)
        
        return Response(result)


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications
    """
    serializer_class = NotificationSerializer
    permission_classes = []  # AllowAny for development
    pagination_class = None  # Disable pagination for notifications
    
    def get_queryset(self):
        """Get notifications for the current user or all users if no user specified"""
        # Automatically create due date notifications when checking notifications
        create_due_date_notifications()
        
        queryset = Notification.objects.select_related('user', 'category', 'submission').all()
        
        # Filter by user if provided
        user_id = self.request.query_params.get('user_id', None)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by read status
        is_read = self.request.query_params.get('is_read', None)
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')
        
        return queryset.order_by('-created_at')
    
    @action(detail=False, methods=['get'], url_path='generate')
    def generate_notifications(self, request):
        """Generate notifications for due dates and approvals"""
        from django.utils import timezone
        from datetime import timedelta
        
        today = timezone.now().date()
        three_days_from_now = today + timedelta(days=3)
        
        notifications_created = 0
        
        # Get all active categories with pending/submitted submissions
        categories = EvidenceCategory.objects.filter(is_active=True).select_related(
            'assignee', 'approver'
        ).prefetch_related('submissions')
        
        for category in categories:
            # Get current submission
            current_submission = category.submissions.filter(
                status__in=['PENDING', 'SUBMITTED', 'UNDER_REVIEW']
            ).order_by('-due_date').first()
            
            if not current_submission:
                continue
            
            # Notifications for assignees - due dates
            if category.assignee:
                days_until_due = (current_submission.due_date - today).days
                
                if days_until_due < 0:
                    # Overdue
                    notification, created = Notification.objects.get_or_create(
                        user=category.assignee,
                        notification_type='OVERDUE',
                        category=category,
                        submission=current_submission,
                        defaults={
                            'title': f'Overdue: {category.name}',
                            'message': f'The submission for {category.name} is overdue by {abs(days_until_due)} day(s).',
                        }
                    )
                    if created:
                        notifications_created += 1
                elif days_until_due <= 3:
                    # Due soon
                    notification, created = Notification.objects.get_or_create(
                        user=category.assignee,
                        notification_type='DUE_SOON',
                        category=category,
                        submission=current_submission,
                        defaults={
                            'title': f'Due Soon: {category.name}',
                            'message': f'The submission for {category.name} is due in {days_until_due} day(s).',
                        }
                    )
                    if created:
                        notifications_created += 1
            
            # Notifications for approvers - pending approvals
            if category.approver and current_submission.status in ['SUBMITTED', 'UNDER_REVIEW']:
                notification, created = Notification.objects.get_or_create(
                    user=category.approver,
                    notification_type='PENDING_APPROVAL',
                    category=category,
                    submission=current_submission,
                    defaults={
                        'title': f'Pending Approval: {category.name}',
                        'message': f'Submission for {category.name} is waiting for your approval.',
                    }
                )
                if created:
                    notifications_created += 1
        
        return Response({
            'message': f'Generated {notifications_created} notifications',
            'notifications_created': notifications_created
        })
    
    @action(detail=True, methods=['post'], url_path='mark-read')
    def mark_read(self, request, pk=None):
        """Mark a notification as read"""
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'message': 'Notification marked as read'})
    
    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        """Mark all notifications as read for a user"""
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=400)
        
        updated = Notification.objects.filter(user_id=user_id, is_read=False).update(is_read=True)
        return Response({'message': f'Marked {updated} notifications as read'})
    
    @action(detail=False, methods=['get'], url_path='unread-count')
    def unread_count(self, request):
        """Get count of unread notifications for a user"""
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=400)
        
        count = Notification.objects.filter(user_id=user_id, is_read=False).count()
        return Response({'unread_count': count})
